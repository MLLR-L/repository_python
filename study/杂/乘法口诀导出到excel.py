# !python3！
# 输入数字以创建乘法法则表格

import openpyxl, os, sys
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter, column_index_from_string

print('请输入一个数字来构建一个乘法口诀表'.center(50,'-'))
if len(sys.argv) > 1:
      excel_input = int(sys.argv[1])
      print(excel_input)
else:
      excel_input = int(input('Enter a number：'))
wb = openpyxl.Workbook()
sheet = wb.active

font_bold = Font(size = 14,bold=True)# 定义字体与加粗样式

# 循环写入表头
for i in range(1, excel_input + 1):
      sheet['A' + str(i+1)] = i
      sheet['A' + str(i+1)].font = font_bold
      column_border = get_column_letter(i+1)
      sheet[column_border + '1'] = i
      sheet[column_border + '1'].font = font_bold


# 写入乘法公式
for i in range(1, excel_input + 1):
      for n in range(1, excel_input + 1):
            column_num = get_column_letter(n+1)
            sheet[str(column_num)+str(i+1)] = i * n
            # print('第',i,'行，第',n,'个')
            #在第一个i里面，n循环写入指定的次数，i也循环直到指定的次数

# 保存
wb.save('乘法口诀表.xlsx')      
wb.close()
open_excel = input('已保存乘法法则.xlsx\n是否要打开？y/回车')
if open_excel == 'y':
      os.startfile('乘法口诀表.xlsx')

      

      
