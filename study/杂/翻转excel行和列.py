# python3!
# 反转excel表格内的内容

import openpyxl, pprint
from openpyxl.utils import get_column_letter, column_index_from_string

wb_last = openpyxl.load_workbook('被反转数据.xlsx')

sheet = wb_last.active
list_last = []

max_row_last = sheet.max_row
max_column_last = sheet.max_column

print('正在读取数据...')

wb_new = openpyxl.Workbook()
sheet_new = wb_new.active
print('正在写入数据')
if max_row_last >max_column_last:
      
      #横向排版
      for i in range(1, max_row_last+1):
            list_last.append([sheet['A' + str(i)].value,sheet['B' + str(i)].value])
      # pprint.pprint(list_last)
      # 写入
      for i in range(int(len(list_last))):
            sheet_new[str(get_column_letter(i+1))+'1'] = str(list_last[i][0])
            sheet_new[str(get_column_letter(i+1))+'2'] = str(list_last[i][1])
else:
      #竖向排版
      for i in range(1, max_column_last+1):
            list_last.append([sheet[str(get_column_letter(i)+'1')].value, sheet[str(get_column_letter(i))+'2'].value])

      #pprint.pprint(list_last)
      
      for i in range(int(len(list_last))):
            sheet_new['A'+str(i+1)] = str(list_last[i][0])
            sheet_new['B'+str(i+1)] = str(list_last[i][1])

File_save='反转后数据.xlsx'
wb_new.save(File_save)
wb_last.close()
wb_new.close()
print('存储名为：'+File_save)
"""
pairs=[[1,2],[3,2],[4,7]]
items =[[x, y]for (y, x) in pairs]
items
[[2, 1], [2, 3], [7, 4]]
"""
